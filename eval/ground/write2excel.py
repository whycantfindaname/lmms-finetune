import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import argparse

# Argument parsing
parser = argparse.ArgumentParser()
parser.add_argument('--input_file', type=str, default='./overall_metrics_results.json', help='path to the input JSON file')
parser.add_argument('--output_file', type=str, default='./overall_metrics_results.xlsx', help='path to the output Excel file')
parser.add_argument('--main_header_label', type=str, default='qwenvl-ft(onetask, 5 epoch)', help='the main header label for the Excel file')
args = parser.parse_args()

# Define the desired order for the categories
desired_order = [
    "Low Clarity", "Blocking Artifacts", "Meaningless Solid Color", "Edge Aliasing Effect",
    "Motion Blur", "Out Of Focus Blur", "Overexposure", "Noise", "Underexposure",
    "Excessive Darkness", "Interlaced Scanning", "Edge Ringing Effect", "Banding Effect", "Moiré Pattern"
]

def load_json_data(input_file):
    """Load JSON data from the specified file."""
    with open(input_file, 'r', encoding='utf-8') as file:
        return json.load(file)

def create_metric_df(data, metric, summary_key=None):
    """
    Create a DataFrame for a specific metric (IoU, AP, or F1).
    Adds summary rows based on the metric type (overall IoU, mAP, average_F1).
    """
    rows = []
    for degradation in desired_order:
        row = {metric: degradation}
        for method in methods:
            value = data[method][metric].get(degradation)
            row[method] = f"{value:.4f}" if value is not None else None
        rows.append(row)

    # Add summary rows for overall IoU, mAP, or average F1
    if metric == 'IoU':
        summary_label = 'Overall IoU between tasks'
        summary_value = data.get('overall_iou_between_models')
        summary_row = {
            metric: summary_label,
            methods[0]: f"{summary_value:.4f}" if summary_value is not None else None,
            methods[1]: f"{summary_value:.4f}" if summary_value is not None else None
        }
        rows.append(summary_row)
    elif summary_key:
        summary_label = summary_key.capitalize()
        summary_row = {metric: summary_label}
        for method in methods:
            summary_value = data[method].get(summary_key)
            summary_row[method] = f"{summary_value:.4f}" if summary_value is not None else None
        rows.append(summary_row)

    return pd.DataFrame(rows)

def save_to_excel_with_custom_header(iou_df, ap_df, f1_df, output_file):
    """
    Save data to an Excel file with each metric on a separate sheet and apply custom header formatting.
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for df, sheet_name in zip([iou_df, ap_df, f1_df], ['IoU', 'AP', 'F1']):
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2, header=False)

    wb = load_workbook(output_file)
    
    for sheet_name in ['IoU', 'AP', 'F1']:
        ws = wb[sheet_name]
        
        # Merge cells for the metric name in A1:A2
        ws["A1"].value = sheet_name
        ws["A1"].font = Font(color="FF0000", bold=True)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:A2")

        # Set the main header label across B1:C1
        ws["B1"].value = args.main_header_label
        ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)

        # Set method names in B2 and C2
        ws["B2"].value = methods[0]
        ws["C2"].value = methods[1]
        ws["B2"].alignment = Alignment(horizontal="center")
        ws["C2"].alignment = Alignment(horizontal="center")

        # Center align data cells
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=3, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    wb.save(output_file)

def format_summary_rows(output_file, sheet_name):
    """
    Format the summary rows by bolding and center-aligning the text.
    """
    wb = load_workbook(output_file)
    ws = wb[sheet_name]
    last_row = ws.max_row
    # Format the summary row
    for cell in ws[f"A{last_row}:C{last_row}"][0]:
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(bold=True)
    wb.save(output_file)

if __name__ == "__main__":
    import os
    if os.path.exists(args.output_file):
        os.remove(args.output_file)

    # Define method labels
    methods = ['pred_vis', 'pred_cap']

    # Load data from JSON
    data = load_json_data(args.input_file)
    
    # Create DataFrames for IoU, AP, and F1 with appropriate summary rows
    iou_df = create_metric_df(data, 'IoU')
    ap_df = create_metric_df(data, 'AP', summary_key='mAP')
    f1_df = create_metric_df(data, 'F1', summary_key='average_F1')
    
    # Save to Excel with custom headers for IoU, AP, and F1
    save_to_excel_with_custom_header(iou_df, ap_df, f1_df, args.output_file)
    
    # Format summary rows in each sheet
    format_summary_rows(args.output_file, 'IoU')
    format_summary_rows(args.output_file, 'AP')
    format_summary_rows(args.output_file, 'F1')
    
    print(f"Data successfully written to {args.output_file} with formatted headers and formatted summary rows.")
